from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Precios catálogo"

FONT_NAME = "Arial"
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", start_color="8F5A45", end_color="8F5A45")
normal_font = Font(name=FONT_NAME, size=10)
bold_font = Font(name=FONT_NAME, bold=True, size=10)
blue_font = Font(name=FONT_NAME, size=10, color="0000FF")
yellow_fill = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
thin = Side(style="thin", color="D9D0C3")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

CATEGORY_LABELS = {
    "colares": "Collar / Gargantilla",
    "pulseiras": "Pulsera",
    "pingentes": "Colgante",
    "brincos": "Arete",
    "aneis": "Anillo",
}

# id, categoria, nombre_es, nombre_pt, costo_brl
ITEMS = [
    ("gargantilha-riviera", "colares", "Gargantilla Riviera con Circonias", "Gargantilha Riviera com Zircônias", 90.92),
    ("colar-elo-portugues", "colares", "Collar Eslabón Portugués Corazón e Infinito", "Colar Elo Português Coração e Infinito", 70.12),
    ("gargantilha-elos-circulares", "colares", "Gargantilla Eslabones Circulares Entrelazados", "Gargantilha Elos Circulares Entrelaçados", 96.04),
    ("colar-cadeado-arvore-vida", "colares", "Collar Candado Árbol de la Vida", "Colar Cadeado Árvore da Vida", 60.28),
    ("colar-cadeado-coracao", "colares", "Collar Candado Corazón", "Colar Cadeado Coração", 47.40),
    ("gargantilha-coracao-cravejado", "colares", "Gargantilla Corazón Engastado", "Gargantilha Coração Cravejado", 61.77),
    ("gargantilha-losangos", "colares", "Gargantilla Dos Rombos Engastados", "Gargantilha Dois Losangos Cravejados", 57.19),
    ("gargantilha-circulos-entrelacados", "colares", "Gargantilla Círculos Entrelazados", "Gargantilha Círculos Entrelaçados", 60.49),
    ("gargantilha-madreperola-sol", "colares", "Gargantilla Medalla de Nácar y Sol", "Gargantilha Medalha de Madrepérola e Sol", 67.31),
    ("gargantilha-madreperola-oval", "colares", "Gargantilla Nácar Ovalado", "Gargantilha Madrepérola Oval", 76.01),
    ("gargantilha-madreperola-natural", "colares", "Gargantilla Medalla de Nácar Natural", "Gargantilha Medalha de Madrepérola Natural", 35.27),
    ("gargantilha-madreperola-acabamento", "colares", "Gargantilla Nácar Acabado Rodio", "Gargantilha Madrepérola Acabamento Ródio", 69.48),
    ("gargantilha-ponto-de-luz", "colares", "Gargantilla Punto de Luz", "Gargantilha Ponto de Luz", 45.75),
    ("pulseira-elos-ovais", "pulseiras", "Pulsera Eslabones Ovalados Engastados", "Pulseira Elos Ovais Cravejados", 51.27),
    ("pulseira-entrelacada-zirconias", "pulseiras", "Pulsera Entrelazada con Circonias", "Pulseira Entrelaçada com Zircônias", 38.46),
    ("pulseira-infinito", "pulseiras", "Pulsera Infinito Engastado", "Pulseira Infinito Cravejado", 57.93),
    ("pingente-arvore-da-vida", "pingentes", "Colgante Árbol de la Vida", "Pingente Árvore da Vida", 38.89),
    ("brinco-solitario-quadrado", "brincos", "Arete Solitario Circonia Cuadrada", "Brinco Solitário Zircônia Quadrada", 45.50),
    ("brinco-argola-cravejado", "brincos", "Arete Aro Engastado", "Brinco Argola Cravejada", 59.68),
    ("brinco-argola-click", "brincos", "Arete Aro Click con Circonia", "Brinco Argola Click com Zircônia", 55.27),
    ("anel-fios-ondulados", "aneis", "Anillo Hilos Ondulados con Circonias", "Anel Fios Ondulados com Zircônias", 64.93),
    ("brinco-gota-zirconias", "brincos", "Arete Gota con Circonias", "Brinco Gota com Zircônias", 32.58),
    ("anel-ajustavel-3-linhas", "aneis", "Anillo Ajustable 3 Líneas de Circonias", "Anel Ajustável 3 Linhas de Zircônias", None),
    ("anel-moissanite-oval", "aneis", "Anillo Moissanita Ovalada", "Anel Moissanite Oval", 117.78),
    ("anel-solitario-moissanite-quadrada", "aneis", "Anillo Solitario Moissanita Cuadrada", "Anel Solitário Moissanite Quadrada", 84.80),
    ("anel-solitario-moissanite-gota", "aneis", "Anillo Solitario Moissanita Talla Gota", "Anel Solitário Moissanite Lapidação Gota", 128.10),
]

# --- Bloque de supuestos ---
ws["A1"] = "Divisor de margen (precio de venta = costo / divisor)"
ws["A1"].font = bold_font
ws["C1"] = 0.4
ws["C1"].font = blue_font
ws["C1"].fill = yellow_fill
ws["C1"].number_format = "0.00"
ws["D1"] = "Costo 40% del precio de venta = margen bruto del 60%. Edita C1 para cambiar el margen."
ws["D1"].font = Font(name=FONT_NAME, size=9, italic=True, color="6B5647")

ws["A2"] = "Tipo de cambio USD/BRL"
ws["A2"].font = bold_font
ws["C2"] = 5.18
ws["C2"].font = blue_font
ws["C2"].fill = yellow_fill
ws["C2"].number_format = "0.00"
ws["D2"] = "1 USD = C2 BRL. Debe coincidir con CONFIG.usdBrlRate en js/main.js. Fuente: exchangerate-api.com, 01/07/2026."
ws["D2"].font = Font(name=FONT_NAME, size=9, italic=True, color="6B5647")

header_row = 4
headers = ["Código", "Categoría", "Nombre (ES)", "Nombre (PT)", "Costo (R$)", "Precio de venta (R$)", "Precio de venta (US$)"]
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

first_data_row = header_row + 1
for i, (pid, cat, name_es, name_pt, cost) in enumerate(ITEMS):
    r = first_data_row + i
    ws.cell(row=r, column=1, value=pid).font = normal_font
    ws.cell(row=r, column=2, value=CATEGORY_LABELS[cat]).font = normal_font
    ws.cell(row=r, column=3, value=name_es).font = normal_font
    ws.cell(row=r, column=4, value=name_pt).font = normal_font

    cost_cell = ws.cell(row=r, column=5, value=cost)
    cost_cell.font = blue_font
    cost_cell.number_format = '#,##0.00;(#,##0.00);"-"'

    sale_cell = ws.cell(row=r, column=6)
    sale_cell.value = f"=IF(E{r}=\"\",\"\",E{r}/$C$1)"
    sale_cell.font = bold_font
    sale_cell.number_format = '#,##0.00;(#,##0.00);"-"'

    sale_usd_cell = ws.cell(row=r, column=7)
    sale_usd_cell.value = f"=IF(E{r}=\"\",\"\",F{r}/$C$2)"
    sale_usd_cell.font = normal_font
    sale_usd_cell.number_format = '#,##0.00;(#,##0.00);"-"'

    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border
        if c not in (3, 4):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

widths = [30, 20, 40, 40, 12, 16, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save("precios_catalogo.xlsx")
print("done")
